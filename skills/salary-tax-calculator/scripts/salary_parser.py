#!/usr/bin/env python3
"""
工资表解析器
支持 Excel (.xlsx/.xls) 和 CSV 格式
自动识别列结构，标准化输出
"""

import json
import argparse
import sys
import os
import csv
from pathlib import Path

SKILL_DIR = Path(__file__).resolve().parent.parent

# 列名映射：各种可能的列名 → 标准字段
COLUMN_MAP = {
    "name": ["姓名", "name", "员工姓名", "员工", "员工名称", "人员", "名字"],
    "gross": ["税前工资", "gross", "应发工资", "应发", "基本工资", "税前", "gross_pay", "工资", "月薪", "薪资"],
    "si_base": ["社保基数", "si_base", "社保缴费基数", "缴费基数", "养老保险基数"],
    "fund_base": ["公积金基数", "fund_base", "公积金缴费基数", "住房公积金基数"],
    "fund_rate": ["公积金比例", "fund_rate", "公积金缴费比例"],
    "deductions": ["专项扣除", "deductions", "专项附加扣除", "special_deductions", "附加扣除"],
    "month": ["月份", "month", "所属月份"],
    "department": ["部门", "department", "所属部门", "组织"],
    "employee_id": ["工号", "id", "employee_id", "员工编号", "编号"],
    "bank_account": ["银行卡号", "bank_account", "银行账号", "工资卡"],
    "social_insurance": ["社保", "social_insurance", "五险", "社保合计"],
    "housing_fund": ["公积金", "housing_fund", "住房公积金", "公积金合计"],
    "tax": ["个税", "tax", "个人所得税", "代扣个税"],
    "net_pay": ["实发工资", "net_pay", "实发", "税后工资", "到手工资", "net"],
    "bonus": ["奖金", "bonus", "绩效", "绩效奖金", "年终奖"],
    "allowance": ["津贴", "allowance", "补贴", "补助"],
    "overtime": ["加班费", "overtime", "加班工资"],
}


def normalize_header(header):
    """标准化表头，返回 {标准字段: 列索引}"""
    mapping = {}
    for idx, col_name in enumerate(header):
        col_lower = col_name.strip().lower()
        for std_field, aliases in COLUMN_MAP.items():
            if col_lower in [a.lower() for a in aliases]:
                mapping[std_field] = idx
                break
    return mapping


def detect_format(filepath):
    """检测文件格式"""
    ext = Path(filepath).suffix.lower()
    if ext in [".xlsx", ".xls"]:
        return "excel"
    elif ext == ".csv":
        return "csv"
    elif ext == ".json":
        return "json"
    else:
        raise ValueError(f"不支持的文件格式: {ext}，支持 .xlsx/.xls/.csv/.json")


def parse_excel(filepath):
    """解析 Excel 文件"""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("请安装 openpyxl: pip install openpyxl")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], None

    header = [str(c) if c is not None else "" for c in rows[0]]
    mapping = normalize_header(header)

    employees = []
    for row in rows[1:]:
        emp = {}
        for field, idx in mapping.items():
            val = row[idx] if idx < len(row) else None
            emp[field] = val
        # 只有当至少有一个关键字段有值时才加入
        if emp.get("name") or emp.get("gross") or emp.get("employee_id"):
            employees.append(emp)

    return employees, mapping


def parse_csv(filepath):
    """解析 CSV 文件"""
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        return [], None

    header = rows[0]
    mapping = normalize_header(header)

    employees = []
    for row in rows[1:]:
        emp = {}
        for field, idx in mapping.items():
            val = row[idx] if idx < len(row) else ""
            emp[field] = val
        if emp.get("name") or emp.get("gross") or emp.get("employee_id"):
            employees.append(emp)

    return employees, mapping


def parse_json(filepath):
    """解析 JSON 文件（已是标准化格式）"""
    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)

    if isinstance(data, list):
        return data, None
    elif isinstance(data, dict) and "employees" in data:
        return data["employees"], None
    else:
        return [data], None


def normalize_employee(emp):
    """标准化员工数据：转换类型、补默认值"""
    result = {}
    # 姓名
    result["name"] = str(emp.get("name", emp.get("employee_id", ""))).strip()

    # 税前工资
    gross = emp.get("gross", emp.get("gross_pay", 0))
    result["gross"] = safe_float(gross)

    # 社保基数
    si_base = emp.get("si_base", emp.get("social_insurance_base"))
    result["si_base"] = safe_float(si_base) if si_base else None

    # 公积金基数
    fund_base = emp.get("fund_base", emp.get("housing_fund_base"))
    result["fund_base"] = safe_float(fund_base) if fund_base else None

    # 公积金比例
    fund_rate = emp.get("fund_rate")
    result["fund_rate"] = safe_float(fund_rate) if fund_rate else None

    # 专项扣除
    deductions = emp.get("deductions", emp.get("special_deductions", 0))
    result["special_deductions"] = safe_float(deductions)

    # 月份
    month = emp.get("month", 1)
    result["month"] = int(safe_float(month) or 1)

    # 其他字段保留
    for extra in ["department", "employee_id", "bank_account", "bonus", "allowance", "overtime"]:
        if emp.get(extra):
            result[extra] = emp[extra]

    return result


def safe_float(value):
    """安全转换为浮点数"""
    if value is None:
        return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0


def parse_salary_file(filepath):
    """主入口：解析工资文件"""
    fmt = detect_format(filepath)

    if fmt == "excel":
        employees, mapping = parse_excel(filepath)
    elif fmt == "csv":
        employees, mapping = parse_csv(filepath)
    elif fmt == "json":
        employees, mapping = parse_json(filepath)
    else:
        raise ValueError(f"不支持的文件格式: {fmt}")

    # 标准化
    normalized = [normalize_employee(e) for e in employees]

    # 过滤有效行（有姓名且有工资）
    valid = [e for e in normalized if e["name"] and e["gross"] > 0]

    return {
        "file": os.path.basename(filepath),
        "format": fmt,
        "total_rows": len(employees),
        "valid_employees": len(valid),
        "column_mapping": {k: v for k, v in (mapping or {}).items()},
        "employees": valid
    }


def main():
    parser = argparse.ArgumentParser(description="工资表解析器")
    parser.add_argument("file", nargs="?", help="工资表文件路径（.xlsx/.xls/.csv/.json）")
    parser.add_argument("--output", "-o", help="输出JSON文件路径")
    parser.add_argument("--template", "-t", action="store_true", help="生成Excel模板")
    parser.add_argument("--city", default="北京", help="批量计算城市")

    args = parser.parse_args()

    if args.template:
        generate_template(args.output or str(SKILL_DIR / "data" / "salary_template.xlsx"))
        return

    if not args.file:
        print("请提供工资表文件路径，或使用 -t 生成模板")
        sys.exit(1)

    result = parse_salary_file(args.file)
    output = json.dumps(result, ensure_ascii=False, indent=2)

    if args.output:
        output_path = Path(args.output)
        if not output_path.is_absolute():
            output_path = SKILL_DIR / "data" / output_path
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(output)
        print(f"解析结果已保存至: {output_path}")
    else:
        print(output)


def generate_template(output_path):
    """生成工资表 Excel 模板"""
    try:
        import openpyxl
    except ImportError:
        print("请安装 openpyxl: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "工资表"

    # 表头
    headers = [
        "姓名", "工号", "部门", "税前工资", "社保基数", "公积金基数",
        "公积金比例", "专项扣除", "月份", "奖金", "津贴", "加班费", "银行卡号"
    ]
    # 样式
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center

    # 示例数据
    examples = [
        ["张三", "EMP001", "技术部", 25000, "", "", "", 3000, 1, "", "", "", ""],
        ["李四", "EMP002", "市场部", 18000, 15000, 15000, 0.07, 2000, 1, 3000, 500, "", ""],
        ["王五", "EMP003", "财务部", 35000, 30000, 30000, 0.12, 4000, 1, "", "", "", ""],
    ]
    for row_idx, data in enumerate(examples, 2):
        for col_idx, val in enumerate(data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # 列宽
    widths = [12, 12, 10, 12, 12, 12, 12, 12, 8, 12, 12, 12, 20]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    # 添加说明 sheet
    ws2 = wb.create_sheet("填写说明")
    instructions = [
        ["字段", "说明", "是否必填", "示例"],
        ["姓名", "员工姓名", "是", "张三"],
        ["工号", "员工工号", "否", "EMP001"],
        ["部门", "所属部门", "否", "技术部"],
        ["税前工资", "应发工资总额", "是", "25000"],
        ["社保基数", "社保缴费基数，空=按实际工资", "否", "25000"],
        ["公积金基数", "公积金缴费基数，空=按实际工资", "否", "25000"],
        ["公积金比例", "如填0.07表示7%，空=按城市默认", "否", "0.07"],
        ["专项扣除", "专项附加扣除总额（子女教育+房贷+赡养老人等）", "否", "3000"],
        ["月份", "所属月份 1-12", "否", "1"],
        ["奖金", "绩效/项目奖金", "否", "3000"],
        ["津贴", "补贴/补助", "否", "500"],
        ["加班费", "加班工资", "否", ""],
        ["银行卡号", "工资卡号", "否", ""],
    ]
    for row_idx, data in enumerate(instructions, 1):
        for col_idx, val in enumerate(data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            if row_idx == 1:
                cell.font = Font(bold=True)

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 45
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 15

    wb.save(output_path)
    print(f"模板已生成: {output_path}")


if __name__ == "__main__":
    main()
